import discord
from discord.ext import commands
from openpyxl import load_workbook
from core.classes import Cog_Extension
import pygsheets
import datetime
import random
import json

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

def count(c):
    cwb = load_workbook('cmdcount.xlsx')
    cws = cwb.active
    if cws['A1'].value != None:
        a = int(cws['A1'].value)
        for i in range(a):
            if str(cws['B' + str(i+1)].value) == str(c.author.id):
                cws['C' + str(i+1)].value = int(cws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    cws['A1'].value = int(cws['A1'].value) + 1
                    cws['B' + str(i+2)].value = str(c.author.id)
                    cws['C' + str(i+2)].value = 1
    else:
        cws['A1'].value = 1
        cws['B1'].value = str(c.author.id)
        cws['C1'].value = 1

    cwb.save('cmdcount.xlsx')
    cwb.close()

def pickcount(c):
    wb = load_workbook('item.xlsx')
    ws = wb.active
    if iws['A1'].value != None:
        a = int(iws['A1'].value)
        for i in range(1, a + 1):
            if str(iws['A' + str(i+1)].value) == str(c.author.id):
                iws['I' + str(i+1)].value += 1
                break

    iwb.save('item.xlsx')

def mine():
    r = random.randint(1, 100)
    num = 0
    I = ['B', 'C', 'D', 'E', 'F', 'G']
    if r < 25:
        it = random.choice(I[:3])
        if it == 'B':
            num = 100
        else:
            num = 5 * 10**(2 - (ord(it) - ord('B')))
    elif r < 45:
        it = random.choice(I[:4])
        if it == 'B':
            num = 200
        else:
            num = 1 * 10**(3 - (ord(it) - ord('B')))
    elif r < 60:
        it = random.choice(I[:3])
        if it == 'B':
            num = 500
        else:
            num = 25 * 10**(2 - (ord(it) - ord('B')))
    elif r < 72:
        it = random.choice(I[:4])
        if it == 'B':
            num = 1000
        else:
            num = 5 * 10**(3 - (ord(it) - ord('B')))
    elif r < 82:
        it = random.choice(I[:3])
        if it == 'B':
            num = 1500
        else:
            num = 75 * 10**(2 - (ord(it) - ord('B')))
    elif r < 89:
        it = random.choice(I[:5])
        if it == 'B':
            num = 2000
        else:
            num = 1 * 10**(4 - (ord(it) - ord('B')))
    elif r < 94:
        it = random.choice(I[:4])
        if it == 'B':
            num = 5000
        else:
            num = 25 * 10**(3 - (ord(it) - ord('B')))
    elif r < 97:
        it = random.choice(I[:5])
        if it == 'B':
            num = 10000
        else:
            num = 5 * 10**(4 - (ord(it) - ord('B')))
    elif r < 99:
        it = random.choice(I[:4])
        if it == 'B':
            num = 15000
        else:
            num = 75 * 10**(3 - (ord(it) - ord('B')))
    else:
        it = random.choice(I)
        if it == 'B':
            num = 20000
        else:
            num = 1 * 10**(5 - (ord(it) - ord('B')))

    return it, num

iwb = load_workbook('item.xlsx')
iws = iwb.active
gc = pygsheets.authorize(service_account_file='gcat-project-42105335-d7a31cee2783.json')
survey_url = 'https://docs.google.com/spreadsheets/d/1L8jP0oFsWRd0fDrPqlgs6vGJLwMaYd5S35-l7Quoc4U/edit#gid=0'
sh = gc.open_by_url(survey_url)
ws = sh.worksheet_by_title('sheet1')
k = -1

class Fun(Cog_Extension):
    global iwb, iws, sh, ws

    @commands.command()
    @commands.is_owner()
    async def system_give(self, ctx, amount: int, typee, name: discord.Member = None):
        count(ctx)
        if name == None:
            if iws['A1'].value != None:
                a = int(iws['A1'].value)
                for i in range(1, a + 1):
                    iws[typee + str(i+1)].value += amount
            else:
                await ctx.send('can\'t find any user')

            await ctx.send(f'Give EVERYONE **{amount} {iws[typee + "1"].value}** meow!!')
        else:
            if iws['A1'].value != None:
                a = int(iws['A1'].value)
                for i in range(1, a + 1):
                    if str(iws['A' + str(i+1)].value) == str(name.id):
                        iws[typee + str(i+1)].value += amount
                        await ctx.send(f'Give **{name} {amount} {iws[typee + "1"].value}** meow!!')
                        break
                    else:
                        if i == a:
                            await ctx.send('Can\'t find the user meow')
                else:
                    await ctx.send('can\'t find any user')
            
        iwb.save('item.xlsx')

    @commands.command()
    @commands.is_owner()
    async def system_bag(self, ctx, name: discord.Member):
        count(ctx)
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(name.id):
                    I = ['B', 'C', 'D', 'E', 'F', 'G', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
                    embed=discord.Embed(title=f'{name}\'s backpack',color=0xffe26f)

                    for it in I:
                        embed.add_field(name=f':small_orange_diamond: **{iws[it + "1"].value}**', value=f'{iws[it + str(i+1)].value}', inline=True)
                        
                    break
                else:
                    if i == a:
                        await ctx.send(f'{name.id} don\'t have any property')
        else:
            await ctx.send('can\'t find any user')

        await ctx.send(embed=embed)
        iwb.save('item.xlsx')

    @commands.command()
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def rank(self, ctx, name: discord.Member = None):
        count(ctx)
        if iws['A1'].value != None:
            if name == None:
                name = ctx.author
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(name.id):
                    I = ['B', 'C', 'D', 'E', 'F', 'G']
                    pro = 0
                    val = 1
                    rk = 1
                    for x in I:
                        pro += iws[x + str(i+1)].value*val
                        if val == 1:
                            val = 2
                        else:
                            val*=10
                    for j in range(1, a + 1):
                        tpro = 0
                        val = 1
                        for x in I:
                            tpro += iws[x + str(j+1)].value*val
                            if val == 1:
                                val = 2
                            else:
                                val*=10
                        if tpro > pro:
                            rk+=1
                        
                    embed=discord.Embed(title=f'{name} is in **#{rk}** meow!!', description=f'Property = **{pro} <:Gcoin:736650744861556749>**', color=0xffe26f)    
                    break
                else:
                    if i == a:
                        await ctx.send('You are not in rank meow!')
                        return
        else:
            await ctx.send('can\'t find any user')
            return

        await ctx.send(embed=embed)
        iwb.save('item.xlsx')


    @commands.command()
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def buy(self, ctx, amount: int, obj):
        count(ctx)
        a = 0
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(ctx.author.id):
                    a = i+1
                    if iws['B' + str(i+1)].value < 0:
                        await ctx.send(f':x: You are in debt meow!')
                        return
                    break
                else:
                    if i == a:
                        await ctx.send('You don\'t have an account(enter .pick first meow!)')
        else:
            await ctx.send('can\'t find any user')
        if obj == 'Copper':
            if iws['B' + str(a)].value - amount*3 >= 0:
                iws['B' + str(a)].value -= amount*3
                iws['C' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Copper** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'Silver':
            if iws['B' + str(a)].value - amount*30 >= 0:
                iws['B' + str(a)].value -= amount*30
                iws['D' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Silver** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'Gold':
            if iws['B' + str(a)].value - amount*400 >= 0:
                iws['B' + str(a)].value -= amount*400
                iws['E' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Gold** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'Diamond':
            if iws['B' + str(a)].value - amount*8000 >= 0:
                iws['B' + str(a)].value -= amount*8000
                iws['F' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Diamond** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'MG':
            if iws['B' + str(a)].value - amount*400000 >= 0:
                iws['B' + str(a)].value -= amount*400000
                iws['G' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Miracle Gem** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'TNT':
            if iws['B' + str(a)].value - amount*500 >= 0:
                iws['B' + str(a)].value -= amount*500
                iws['J' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} TNT** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif  obj == 'Dynamite':
            if iws['B' + str(a)].value - amount*1000 >= 0:
                iws['B' + str(a)].value -= amount*1000
                iws['K' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Dynamite** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'Knife':
            if iws['B' + str(a)].value - 5000 >= 0:
                if iws['L' + str(a)].value == 0:
                    iws['B' + str(a)].value -= 5000
                    iws['L' + str(a)].value += 1
                    await ctx.send(f':ballot_box_with_check: It\'s your **Knife** meow!Thank you for coming meow~')
                else:
                    await ctx.send(f':x: You have owned a Knife meow!')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'DE':
            if iws['B' + str(a)].value - 35000 >= 0:
                if iws['M' + str(a)].value == 0:
                    iws['B' + str(a)].value -= 35000
                    iws['M' + str(a)].value += 1
                    await ctx.send(f':ballot_box_with_check: It\'s your **Desert Eagle** meow!Thank you for coming meow~')
                else:
                    await ctx.send(f':x: You have owned a Desert Eagle meow!')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'MP5':
            if iws['B' + str(a)].value - 50000 >= 0:
                if iws['N' + str(a)].value == 0:
                    iws['B' + str(a)].value -= 50000
                    iws['N' + str(a)].value += 1
                    await ctx.send(f':ballot_box_with_check: It\'s your **MP5** meow!Thank you for coming meow~')
                else:
                    await ctx.send(f':x: You have owned a MP5 meow!')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'Bullet(DE)':
            if iws['B' + str(a)].value - amount*100 >= 0:
                iws['B' + str(a)].value -= amount*100
                iws['O' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Bullet(DE)** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        elif obj == 'Magazine(MP5)':
            if iws['B' + str(a)].value - amount*200 >= 0:
                iws['B' + str(a)].value -= amount*200
                iws['P' + str(a)].value += amount
                await ctx.send(f':ballot_box_with_check: It\'s your **{amount} Magazine(MP5)** meow!Thank you for coming meow~')
            else:
                await ctx.send(f':x: You don\'t have enough money meow!')
        # elif obj == 'PickCD':
        #     if iws['Q' + str(a)].value == 1800:
        #         await ctx.send(':x: The minimum for PickCD is 1800 sec meow!')
        #         return
        #     if iws['B' + str(a)].value - amount*20000 >= 0 and iws['Q' + str(a)].value - amount*100 >= 1800:
        #         iws['B' + str(a)].value -= amount*20000
        #         iws['Q' + str(a)].value -= amount*100
        #         await ctx.send(f':ballot_box_with_check: You cut down the **PickCD {amount} times ({amount*100}sec)** meow!Thank you for coming meow~')
        #     else:
        #         await ctx.send(f':x: You don\'t have enough money or over the purchase limit of PickCD meow!Try to type less amount meow!')
        # elif obj == 'RobCD':
        #     if iws['R' + str(a)].value == 10800:
        #         await ctx.send(':x: The minimum for RobCD is 10800 sec meow!')
        #         return
        #     if iws['B' + str(a)].value - amount*5000 >= 0 and iws['R' + str(a)].value - amount*100 >= 10800:
        #         iws['B' + str(a)].value -= amount*5000
        #         iws['R' + str(a)].value -= amount*100
        #         await ctx.send(f':ballot_box_with_check: You cut down the **RobCD {amount} times ({amount*100}sec)** meow!Thank you for coming meow~')
        #     else:
        #         await ctx.send(f':x: You don\'t have enough money or over the purchase limit of RobCD meow!Try to type less amount meow!')
        iwb.save('item.xlsx')


    @commands.command()
    @commands.cooldown(1, 30, commands.BucketType.user)
    async def shop(self, ctx):
        embed=discord.Embed(title="桔喵Black Market", description="Product List", color=0xffe26f)
        embed.set_thumbnail(url="https://i.imgur.com/NOq5mPb.png")
        embed.add_field(name=":small_orange_diamond:Copper", value="<:Gcoin:736650744861556749> 3", inline=True)
        embed.add_field(name=":small_orange_diamond:Silver", value="<:Gcoin:736650744861556749> 30", inline=True)
        embed.add_field(name=":small_orange_diamond:Gold", value="<:Gcoin:736650744861556749> 400", inline=True)
        embed.add_field(name=":small_orange_diamond:Diamond", value="<:Gcoin:736650744861556749> 8000", inline=True)
        embed.add_field(name=":small_orange_diamond:Miracle Gem", value="<:Gcoin:736650744861556749> 400000", inline=True)
        embed.add_field(name=":small_orange_diamond:TNT", value="<:Gcoin:736650744861556749> 500", inline=True)
        embed.add_field(name=":small_orange_diamond:Dynamite", value="<:Gcoin:736650744861556749> 1000", inline=True)
        embed.add_field(name=":small_orange_diamond:Knife", value="<:Gcoin:736650744861556749> 5000", inline=True)
        embed.add_field(name=":small_orange_diamond:Desert Eagle", value="<:Gcoin:736650744861556749> 35000", inline=True)
        embed.add_field(name=":small_orange_diamond:MP5", value="<:Gcoin:736650744861556749> 50000", inline=True)
        embed.add_field(name=":small_orange_diamond:Bullet(DE)", value="<:Gcoin:736650744861556749> 100", inline=True)
        embed.add_field(name=":small_orange_diamond:Magazine(MP5)", value="<:Gcoin:736650744861556749> 200", inline=True)
        # embed.add_field(name=":small_orange_diamond:PickCD", value="<:Gcoin:736650744861556749> 20000", inline=True)
        # embed.add_field(name=":small_orange_diamond:RobCD", value="<:Gcoin:736650744861556749> 5000", inline=True)
        # embed.add_field(name=":small_orange_diamond:Comming S∞n", value="<:Gcoin:736650744861556749> ??", inline=True)
        embed.set_footer(text="Thank you for coming meow~")
        await ctx.send(embed=embed)
        iwb.save('item.xlsx')

    @commands.command()
    @commands.cooldown(1, 600, commands.BucketType.user)
    async def give(self, ctx, name: discord.Member, amount: int, typee):
        count(ctx)
        if amount < 0:
            await ctx.send(':x: DON\'t try to rob by by typing “.give” command meow!')
            return
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            isfind = False
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(ctx.author.id):
                    for j in range(1, a + 1):
                        if str(iws['A' + str(j+1)].value) == str(name.id):
                            T = {'Gcoin' : 'B', 'Copper' : 'C', 'Silver' : 'D', 'Gold' : 'E', 'Diamond' : 'F', 'MG' : 'G'}
                            goods = T[typee]
                            if iws['A' + str(i+1)].value == str(name.id):
                                await ctx.send(f':x: Don\'t try to give property to yourself meow!')
                                isfind = True
                                break
                            if iws[goods + str(i+1)].value >= amount:
                                iws[goods + str(i+1)].value -= amount
                                iws[goods + str(j+1)].value += amount
                            else:
                                await ctx.send(f':x: **{ctx.author}** don\'t have enough **{typee}** meow!')
                            await ctx.send(f':handshake:  **{ctx.author}** gave **{name}** **{amount} {typee}** meow!!')
                            isfind = True
                            break
                        else:
                            if j == a:
                                isfind = True
                                await ctx.send(f':x: **{name}** didn\'t have an account meow!')
                else:
                    if i == a and not(isfind):
                        await ctx.send(':x: You don\'t have an account(enter .pick first meow!)')
        else:
            await ctx.send('can\'t find any user')

        iwb.save('item.xlsx')

    @commands.command()
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def picktimes(self, ctx):
        count(ctx)
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(ctx.author.id):
                    await ctx.send(f':raised_back_of_hand: **{ctx.author}** has picked {iws["I" + str(i+1)].value} times meow!')
                    break
                else:
                    if i == a:
                        await ctx.send('You don\'t have an account(enter .pick first meow!)')
        else:
            await ctx.send('can\'t find any user')

        iwb.save('item.xlsx')


    @commands.command()
    @commands.cooldown(1, 80000, commands.BucketType.user)
    async def daily(self, ctx):
        count(ctx)
        if ws.get_value('A1') != None:
            a = int(iws['A1'].value)
            L = ws.get_col(1)[:a+1]
            i = 1
            for x in L:
                if x == L[0]:
                    continue
                i+=1
                if str(ws.get_value('A' + str(i))) == str(ctx.author.id):
                    ws.update_value('B' + str(i), int(ws.get_value('B' + str(i))) + 500)
                    await ctx.send(f':moneybag: **{ctx.author}** earned the daily reward meow!')
                    break
                else:
                    if i == a+1:
                        await ctx.send('You don\'t have an account(enter .pick first meow!)')
        else:
            await ctx.send('can\'t find any user')

        iwb.save('item.xlsx')

    @commands.command()
    @commands.cooldown(2, 10800, commands.BucketType.user)
    async def rob(self, ctx, name: discord.Member):
        count(ctx)
        if ws.get_value('A1') != None:
            a = int(ws.get_value('A1'))
            L = ws.get_col(1)[:a+1]
            i, j = 1, 1
            isfind = False
            for x in L:
                if x == L[0]:
                    continue
                i+=1
                # print(f'{x},{i}')
                if str(x) == str(ctx.author.id):
                    for y in L:
                        if y == L[0]:
                            continue
                        j+=1
                        # print(f'{y},{j}')
                        if str(y) == str(name.id):
                            if float(ws.get_value('H' + str(i))) < 20:
                                isProps = 0
                                ws.update_value('H' + str(i), float(ws.get_value('H' + str(i))) + 0.5)
                                await ctx.send(f'<a:hand:732937258868539483> **{ctx.author}**\'s Robbery skills point + 0.5 ({ws.get_value("H" + str(i))}/20)')
                            r = random.randint(1, 100)
                            if r <= float(ws.get_value('H' + str(i))):
                                P = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
                                p = random.choice(P)
                                if int(ws.get_value('N' + str(i))) == 1 and int(ws.get_value('P' + str(i))) > 0:
                                    p = ws.get_value('B' + str(j))
                                    isProps = 1
                                elif int(ws.get_value('M' + str(i))) == 1 and int(ws.get_value('O' + str(i))) > 0:
                                    p = int(ws.get_value('B' + str(j)))//2
                                    isProps = 2
                                elif int(ws.get_value('L' + str(i))) == 1:
                                    p*=2
                                    isProps = 3
                                ws.update_value('B' + str(i), int(ws.get_value('B' + str(i))) + int(p))
                                ws.update_value('B' + str(j), int(ws.get_value('B' + str(j))) - int(p))

                                if isProps == 0:
                                    await ctx.send(f':money_with_wings: **{ctx.author}** robbed the property form **{name}** meow!!({p} <:Gcoin:736650744861556749>)')
                                elif isProps == 1:
                                    await ctx.send(f':money_with_wings: **{ctx.author}** robbed the property form **{name}** meow!!({p} <:Gcoin:736650744861556749>)\n<a:frog_gun:732828139499159625> Because **{ctx.author}** use **MP5** so all the **{name}**\'s Gcoin was taken away meow!')
                                elif isProps == 2:
                                    await ctx.send(f':money_with_wings: **{ctx.author}** robbed the property form **{name}** meow!!({p} <:Gcoin:736650744861556749>)\n<a:frog_gun:732828139499159625> Because **{ctx.author}** use **Desert Eagle** so half the **{name}**\'s Gcoin was taken away meow!')
                                else:
                                    await ctx.send(f':money_with_wings: **{ctx.author}** robbed the property form **{name}** meow!!({p} <:Gcoin:736650744861556749>)\n<a:frog_gun:732828139499159625> Because **{ctx.author}** use **Knife** so **{ctx.author}** robbed 2x <:Gcoin:736650744861556749> meow!')
                                
                            else:
                                ws.update_value('B' + str(i), int(ws.get_value('B' + str(i))) - 200)
                                await ctx.send(f'<a:money:730029539815850045> **{ctx.author}** failed to rob the property form **{name}**\n:police_car: The MeowPolice took your property away meow~(200 <:Gcoin:736650744861556749>)')
                            
                            if int(ws.get_value('N' + str(i))) == 1 and int(ws.get_value('P' + str(i))) > 0:
                                print('dododo')
                                ws.update_value('P' + str(i), int(ws.get_value('P' + str(i))) - 1)
                            elif int(ws.get_value('M' + str(i))) == 1 and int(ws.get_value('O' + str(i))) > 0:
                                ws.update_value('O' + str(i), int(ws.get_value('O' + str(i))) - 1)
                            isfind = True
                            break
                        else:
                            if j == a+1:
                                isfind = True
                                await ctx.send(f':x: **{name}** didn\'t have an account meow!')
                else:
                    if i == (a+1) and not(isfind):
                        await ctx.send(':x: You don\'t have an account(enter .pick first meow!)')
        else:
            await ctx.send('can\'t find any user')

    @commands.command()
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def ibag(self, ctx):
        count(ctx)
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(ctx.author.id):
                    I = ['J', 'K', 'L', 'M', 'N', 'O', 'P']
                    embed=discord.Embed(title=f'{ctx.author}\'s backpack',color=0xffe26f)

                    for it in I:
                        embed.add_field(name=f':small_orange_diamond: **{iws[it + "1"].value}**', value=f'{iws[it + str(i+1)].value}', inline=False)
                        
                    break
                else:
                    if i == a:
                        await ctx.send('You don\'t have any property')
        else:
            await ctx.send('can\'t find any user')

        await ctx.send(embed=embed)
        iwb.save('item.xlsx')

    @commands.command()
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def bag(self, ctx):
        count(ctx)
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(ctx.author.id):
                    I = ['B', 'C', 'D', 'E', 'F', 'G']
                    embed=discord.Embed(title=f'{ctx.author}\'s backpack',color=0xffe26f)

                    for it in I:
                        embed.add_field(name=f':small_orange_diamond: **{iws[it + "1"].value}**', value=f'{iws[it + str(i+1)].value}', inline=False)
                        
                    break
                else:
                    if i == a:
                        await ctx.send('You don\'t have any property')
        else:
            await ctx.send('can\'t find any user')

        await ctx.send(embed=embed)
        iwb.save('item.xlsx')

    @commands.command()
    @commands.cooldown(2, 1800, commands.BucketType.user)
    async def pick(self, ctx):
        count(ctx)
        if ws.get_value('A1') != '':
            a = int(ws.get_value('A1'))
            L = ws.get_col(1)[:a+1]
            i = 1
            for x in L:
                if x == L[0]:
                    continue
                i+=1
                if str(x) == str(ctx.author.id):
                    it, num = mine()
                    if int(ws.get_value('K' + str(i))) > 0:
                        ws.update_value('K' + str(i), int(ws.get_value('K' + str(i))) - 1)
                        num*=4
                        await ctx.send(f':boom: You blasted a HUGE hole and you found **{num}** **{ws.get_value(it + "1")}**!(x4 income)')
                    elif int(ws.get_value('J' + str(i))) > 0:
                        ws.update_value('J' + str(i), int(ws.get_value('J' + str(i))) - 1)
                        num*=2
                        await ctx.send(f':boom: You blasted a BIG hole and you found **{num}** **{ws.get_value(it + "1")}**!(x2 income)')
                    else:
                        await ctx.send(f'You pick up **{num}** **{ws.get_value(it + "1")}**!')
                    ws.update_value(it + str(i), int(ws.get_value(it + str(i))) + num)
                    break
                else:
                    if i == a+1:
                        ws.update_value('A1', a+1)
                        nL = [0] * 16
                        ws.update_row(a+2, nL)
                        ws.update_value('A' + str(i+1), str(ctx.author.id))
                        it, num = mine()
                        ws.update_value(it + str(i+1), int(ws.get_value(it + str(i+1))) + num)
                        await ctx.send(f'You pick up **{num}** **{ws.get_value(it + "1")}**!')
        else:
            ws.update_value('A1', 1)
            nL = [0] * 16
            ws.update_row(2, nL)
            ws.update_value('A2', str(ctx.author.id))
            it, num = mine()
            ws.update_value(it + '2', int(ws.get_value(it + '2')) + num)
            await ctx.send(f'You pick up **{num}** **{ws.get_value(it + "1")}**!')

        pickcount(ctx)

    @commands.group()
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def sell(self, ctx):
        global k
        if iws['A1'].value != None:
            a = int(iws['A1'].value)
            for i in range(1, a + 1):
                if str(iws['A' + str(i+1)].value) == str(ctx.author.id):
                    k = i+1
                    break
                else:
                    if i == a:
                        await ctx.send('You don\'t have any property')
        else:
            await ctx.send('can\'t find any user')
        embed=discord.Embed(title="Sell", color=0xffe26f)

    @sell.command()
    async def Copper(self, ctx, amount = None):
        if amount != None:
            amount = int(amount)
            if amount > iws['C' + str(k)].value:
                await ctx.send(f':x: You DON\'T HAVE so many treasures meow!!')
                return
        else:
            amount = iws['C' + str(k)].value
        iws['B' + str(k)].value += amount*2
        iws['C' + str(k)].value -= amount
        await ctx.send(f'**{amount} Copper** sold successfully meow!')
        iwb.save('item.xlsx')

    @sell.command()
    async def Silver(self, ctx, amount = None):
        if amount != None:
            amount = int(amount)
            if amount > iws['D' + str(k)].value:
                await ctx.send(f':x: You DON\'T HAVE so many treasures meow!!')
                return
        else:
            amount = iws['D' + str(k)].value
        iws['B' + str(k)].value += amount*20
        iws['D' + str(k)].value -= amount
        await ctx.send(f'**{amount} Silver** sold successfully meow!')
        iwb.save('item.xlsx')

    @sell.command()
    async def Gold(self, ctx, amount = None):
        if amount != None:
            amount = int(amount)
            if amount > iws['E' + str(k)].value:
                await ctx.send(f':x: You DON\'T HAVE so many treasures meow!!')
                return
        else:
            amount = iws['E' + str(k)].value
        iws['B' + str(k)].value += amount*200
        iws['E' + str(k)].value -= amount
        await ctx.send(f'**{amount} Gold** sold successfully meow!')
        iwb.save('item.xlsx')

    @sell.command()
    async def Diamond(self, ctx, amount = None):
        if amount != None:
            amount = int(amount)
            if amount > iws['F' + str(k)].value:
                await ctx.send(f':x: You DON\'T HAVE so many treasures meow!!')
                return
        else:
            amount = iws['F' + str(k)].value
        iws['B' + str(k)].value += amount*2000
        iws['F' + str(k)].value -= amount
        await ctx.send(f'**{amount} Diamond** sold successfully meow!')
        iwb.save('item.xlsx')

    @sell.command()
    async def MG(self, ctx, amount = None):
        if amount != None:
            amount = int(amount)
            if amount > iws['G' + str(k)].value:
                await ctx.send(f':x: You DON\'T HAVE so many treasures meow!!')
                return
        else:
            amount = iws['G' + str(k)].value
        iws['B' + str(k)].value += amount*20000
        iws['G' + str(k)].value -= amount
        await ctx.send(f'**{amount} Miracle Gem** sold successfully meow!')
        iwb.save('item.xlsx')


def setup(bot):
    bot.add_cog(Fun(bot))